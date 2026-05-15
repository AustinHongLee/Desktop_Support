from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from pypdf import PdfReader, PdfWriter

SERIAL_HEADERS = ("流水號", "sort", "serial", "序號", "排序", "順序", "seq", "no", "編號", "項次", "number")
DRAWING_NAME_HEADERS = (
    "file_basename",
    "file basename",
    "basename",
    "source_pdf_name",
    "source pdf name",
    "dst_pdf_name",
    "dst pdf name",
    "pdf name",
    "pdf檔名",
    "檔名",
    "圖號",
    "圖名",
    "drawing no",
    "drawing number",
    "dwg no",
    "dwg",
)
LINE_HEADERS = (
    "管線號碼",
    "管線號",
    "管線編號",
    "線號",
    "line",
    "line no",
    "line number",
    "lineno",
    "管號",
    "piping",
    "pipe",
)


@dataclass(frozen=True)
class IsoRecord:
    serial: str
    line_no: str


@dataclass(frozen=True)
class IsoTable:
    sheet_name: str
    header_row_index: int
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True)
class NamingRow:
    source: Path
    page: int
    serial: str
    line_no: str
    new_name: str


def list_iso_sheets(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return ["CSV"]
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl，無法讀取 Excel ISO list。") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    raise ValueError("ISO list 目前支援 .xlsx, .xlsm, .csv")


def read_iso_table(path: Path, *, sheet_name: str | None = None) -> IsoTable:
    rows, resolved_sheet_name = _read_raw_rows(path, sheet_name=sheet_name)
    if not rows:
        raise ValueError("ISO list 沒有可讀取的資料")
    header_index = _find_header_row(rows)
    headers = tuple(_cell_text(cell) or f"欄位{index + 1}" for index, cell in enumerate(rows[header_index]))
    data_rows = tuple(tuple(row) for row in rows[header_index + 1 :])
    return IsoTable(
        sheet_name=resolved_sheet_name,
        header_row_index=header_index,
        headers=headers,
        rows=data_rows,
    )


def guess_iso_columns(headers: tuple[str, ...]) -> tuple[int | None, int | None]:
    normalized = [_normalize_header(header) for header in headers]
    line_col = _first_header_match(normalized, DRAWING_NAME_HEADERS)
    if line_col is None:
        line_col = _first_header_match(normalized, LINE_HEADERS)
    return (
        _first_header_match(normalized, SERIAL_HEADERS),
        line_col,
    )


def records_from_table(table: IsoTable, *, serial_col: int, line_col: int) -> list[IsoRecord]:
    records: list[IsoRecord] = []
    for row in table.rows:
        serial = _cell_text(row[serial_col] if serial_col < len(row) else "")
        line_no = _drawing_name_text(row[line_col] if line_col < len(row) else "")
        if serial and line_no:
            records.append(IsoRecord(serial=serial, line_no=line_no))
    return records


def read_iso_records(
    path: Path,
    *,
    sheet_name: str | None = None,
    serial_col: int | None = None,
    line_col: int | None = None,
) -> list[IsoRecord]:
    table = read_iso_table(path, sheet_name=sheet_name)
    if serial_col is None or line_col is None:
        guessed_serial_col, guessed_line_col = guess_iso_columns(table.headers)
        serial_col = guessed_serial_col if serial_col is None else serial_col
        line_col = guessed_line_col if line_col is None else line_col
    if serial_col is None or line_col is None:
        raise ValueError("找不到 ISO list 欄位，請手動指定「流水號」與「圖號/檔名」欄位")
    return records_from_table(table, serial_col=serial_col, line_col=line_col)


def _legacy_read_iso_records(path: Path) -> list[IsoRecord]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_records(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx_records(path)
    raise ValueError("ISO list 目前支援 .xlsx, .xlsm, .csv")


def build_naming_rows(
    files: list[Path],
    records: list[IsoRecord],
    *,
    pattern: str = "{serial}--{line}.pdf",
) -> list[NamingRow]:
    lookup = build_record_lookup(records)
    rows: list[NamingRow] = []
    for index, source in enumerate(sorted(files, key=natural_pdf_key), start=1):
        serial = str(index)
        record = lookup.get(_serial_key(serial))
        line_no = record.line_no if record else ""
        rows.append(
            NamingRow(
                source=source,
                page=index,
                serial=serial,
                line_no=line_no,
                new_name=format_iso_name(pattern, serial=serial, line=line_no),
            )
        )
    return rows


def build_record_lookup(records: list[IsoRecord]) -> dict[str, IsoRecord]:
    lookup: dict[str, IsoRecord] = {}
    for record in records:
        for key in _serial_keys(record.serial):
            lookup[key] = record
    return lookup


def format_iso_name(pattern: str, *, serial: str, line: str) -> str:
    serial = serial.strip()
    line = _pdf_basename(line)
    if not serial or not line:
        return ""
    name = pattern.replace("{serial}", serial).replace("{line}", line)
    return name if Path(name).suffix else f"{name}.pdf"


def parse_iso_filename(name: str) -> IsoRecord | None:
    match = re.match(r"^(?P<serial>\d+)--(?P<line>.+?)\.pdf$", name, flags=re.IGNORECASE)
    if not match:
        return None
    return IsoRecord(serial=match.group("serial"), line_no=_pdf_basename(match.group("line")))


def natural_pdf_key(path: Path) -> tuple[Any, ...]:
    text = path.stem.lower()
    parts = re.split(r"(\d+)", text)
    return tuple(int(part) if part.isdigit() else part for part in parts)


def split_pdf_to_pages(pdf: Path) -> list[Path]:
    reader = PdfReader(str(pdf))
    page_count = len(reader.pages)
    output_dir = pdf.with_name(f"{pdf.stem}_pages")
    output_dir.mkdir(exist_ok=True)
    width = max(3, len(str(page_count)))
    outputs: list[Path] = []
    for index, page in enumerate(reader.pages, start=1):
        writer = PdfWriter()
        writer.add_page(page)
        output = output_dir / f"{pdf.stem}_p{index:0{width}d}.pdf"
        with output.open("wb") as handle:
            writer.write(handle)
        outputs.append(output)
    return outputs


def _read_csv_records(path: Path) -> list[IsoRecord]:
    rows, _sheet_name = _read_raw_csv_rows(path)
    return _records_from_rows(rows)


def _read_raw_rows(path: Path, *, sheet_name: str | None = None) -> tuple[list[list[Any]], str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_raw_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_raw_xlsx_rows(path, sheet_name=sheet_name)
    raise ValueError("ISO list 目前支援 .xlsx, .xlsm, .csv")


def _read_raw_csv_rows(path: Path) -> tuple[list[list[Any]], str]:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "cp950", "big5"):
        try:
            with path.open("r", newline="", encoding=encoding) as handle:
                rows = list(csv.reader(handle))
            return rows, "CSV"
        except UnicodeDecodeError as exc:
            last_error = exc
    if last_error:
        raise last_error
    return [], "CSV"


def _read_xlsx_records(path: Path) -> list[IsoRecord]:
    rows, _sheet_name = _read_raw_xlsx_rows(path, sheet_name=None)
    return _records_from_rows(rows)


def _read_raw_xlsx_rows(path: Path, *, sheet_name: str | None) -> tuple[list[list[Any]], str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，無法讀取 Excel ISO list。") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"找不到 Sheet：{sheet_name}")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        return rows, sheet.title
    finally:
        workbook.close()


def _records_from_rows(rows: list[list[Any]]) -> list[IsoRecord]:
    if not rows:
        return []
    header_index, serial_col, line_col = _find_columns(rows)
    records: list[IsoRecord] = []
    for row in rows[header_index + 1 :]:
        serial = _cell_text(row[serial_col] if serial_col < len(row) else "")
        line_no = _drawing_name_text(row[line_col] if line_col < len(row) else "")
        if serial and line_no:
            records.append(IsoRecord(serial=serial, line_no=line_no))
    return records


def _find_columns(rows: list[list[Any]]) -> tuple[int, int, int]:
    header_index = _find_header_row(rows)
    normalized = [_normalize_header(_cell_text(cell)) for cell in rows[header_index]]
    serial_col = _first_header_match(normalized, SERIAL_HEADERS)
    line_col = _first_header_match(normalized, DRAWING_NAME_HEADERS)
    if line_col is None:
        line_col = _first_header_match(normalized, LINE_HEADERS)
    if serial_col is not None and line_col is not None:
        return header_index, serial_col, line_col
    raise ValueError("找不到 ISO list 欄位，至少需要「流水號」與「圖號/檔名」")


def _find_header_row(rows: list[list[Any]]) -> int:
    for row_index, row in enumerate(rows[:20]):
        normalized = [_normalize_header(_cell_text(cell)) for cell in row]
        serial_col = _first_header_match(normalized, SERIAL_HEADERS)
        line_col = _first_header_match(normalized, DRAWING_NAME_HEADERS)
        if line_col is None:
            line_col = _first_header_match(normalized, LINE_HEADERS)
        if serial_col is not None or line_col is not None:
            return row_index
    return 0


def _first_header_match(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    for candidate in candidates:
        normalized_candidate = _normalize_header(candidate)
        for index, header in enumerate(headers):
            if header == normalized_candidate:
                return index
    return None


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-\[\]【】()（）.]", "", value).lower()


def _pdf_basename(value: Any) -> str:
    text = _cell_text(value)
    if text.lower().endswith(".pdf"):
        text = text[:-4]
    return text.strip()


def _drawing_name_text(value: Any) -> str:
    text = _pdf_basename(value)
    match = re.match(r"^\d+--(.+)$", text)
    return match.group(1).strip() if match else text


def _serial_keys(serial: str) -> set[str]:
    key = _serial_key(serial)
    keys = {key}
    if key.isdigit():
        keys.add(str(int(key)))
    return keys


def _serial_key(serial: str) -> str:
    return serial.strip()
